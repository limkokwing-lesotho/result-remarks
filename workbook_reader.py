from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich import print

from model import BorderlineObject
from utils.common import convert_list_to_dict, is_number, to_int

PARSER = "html5lib"


def get_marks_cols(sheet: Worksheet):
    courses = {}
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell.value == "Mk":
                col_i = cell.col_idx
                row_i = cell.row
                grade_cell: Cell = sheet.cell(row_i - 2, col_i)
                # courses[grade_cell.value] = grade_cell.column
                courses[grade_cell.column] = grade_cell.value
    return courses


def get_std_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell.value == "StudentID":
                return cell.column
    return 3


def get_remark_col(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell and cell.value and "remark" in str(cell.value).lower():
                return cell.column
    return -1


def get_student_numbers(sheet: Worksheet) -> dict[int, str]:
    data = {}
    std_col = get_std_column(sheet)

    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell.column == std_col:
                if is_number(cell.value):
                    data[cell.row] = cell.value
    return data


def credit_hours_to_marks(credit_hours):
    marks = (credit_hours * 50) / 2
    return marks


def read_excel_marks(sheet: Worksheet):
    marks_dict = get_marks_cols(sheet)
    students = get_student_numbers(sheet)

    results = {}
    for student_col in students:
        student_number = to_int(students[student_col])
        data = []
        for mark_col, course_code in marks_dict.items():
            cell: Cell = sheet.cell(student_col, mark_col)
            mark_value = None
            if is_number(cell.value):
                mark_value = float(cell.value)
                data.append({course_code: mark_value})
        results[student_number] = data

    return convert_list_to_dict(results)


def get_border_line_objects(sheet: Worksheet):
    marks_dict = get_marks_cols(sheet)
    students = get_student_numbers(sheet)

    data = []
    for student_col in students:
        student_number = to_int(students[student_col])
        for mark_col, course_code in marks_dict.items():
            cell: Cell = sheet.cell(student_col, mark_col)
            mark_value = None
            if is_number(cell.value):
                grade: Cell = sheet.cell(student_col, mark_col + 1)
                mark_value = float(cell.value)
                borderline = BorderlineObject(
                    std_no=student_number,
                    std_class=sheet.title,
                    course=course_code,
                    marks=mark_value,
                    grade=grade.value,
                )
                data.append(borderline)

    return data


def combine_cms_excel_marks(excel_marks: dict[int, dict], cms_marks: dict[int, dict]):
    data = {}

    for std, results in cms_marks.items():
        items = {}
        for course, credits in results.items():
            marks = 0
            if credits:
                marks = credit_hours_to_marks(credits)
            # print("std", std)
            # print("course", course)
            # print("credits", credits)
            # print("marks", marks)

            items[course] = marks
        data[std] = items

    for std, results in excel_marks.items():
        items = {}
        for course, credits in results.items():
            data[std][course] = credits
        # data[std] = items

    return data


def generate_remarks(sheet: Worksheet, cms_marks: dict[int, dict]):
    excel_marks = read_excel_marks(sheet)
    students = get_student_numbers(sheet)
    marks_dict = combine_cms_excel_marks(excel_marks, cms_marks)

    # print(marks_dict)

    data = {}
    for student_col in students:
        student_number = to_int(students[student_col])
        repeat = []
        sup = []
        for course_code, marks in marks_dict[student_number].items():
            if is_number(marks):
                mark_value = float(marks)
                if mark_value >= 45 and mark_value < 50:
                    sup.append(course_code)
                elif mark_value < 45:
                    repeat.append(course_code)
            else:
                print(f"{student_number=} {course_code=} {marks=}")

        remarks = "Proceed"
        if len(repeat) >= 3:
            remarks = "Remain in Semester"

        if len(sup) > 0:
            remarks = f"{remarks}, Sup " + ", ".join(sup)
        if len(repeat) > 0:
            remarks = f"{remarks}, Repeat " + ", ".join(repeat)

        data[student_col] = remarks
    return data


def main():
    pass


if __name__ == "__main__":
    print(main())
